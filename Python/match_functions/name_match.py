import pandas as pd
import re
import openpyxl
import pymysql

# pd.set_option("display.max_columns", None)
# pd.set_option("display.max_rows", None)
#
# #read excel file
# name_table_1 = 'PSP_to_Airtable_1.xlsx'
# name_table_2 = 'PSP_to_Airtable_1_copy.xlsx'
# sheet_name_1 = 'Interac'
# sheet_name_2 = 'Interac'
# df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
# df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
# df1_column = 'FI_Name'
# df2_column = 'FI_Name'
# width1 = len(df1.columns)
# width2 = len(df2.columns)
# length1 = len(df1.index)
# length2 = len(df2.index)

# function which look for connection between two columns (take 2 columns and return dictionary with indexes of rows which have connection)
def find_connections(column1, column2, final_score):
    connection_indexes = {}
    for i in range(max(len(column1), len(column2))):
        connection_indexes[i] = []   # for each key we create list of values
        for j in range(len(column2)):
            if column1[i] == column2[j]:
                connection_indexes[i].append([j, final_score])

    return connection_indexes


# function which write connections and scores to new columns to both tables
def write_connections(name_table_1, df1, sheet1, column1, name_table_2, df2, sheet2, column2, final_score, columntype_score_1, columntype_score_2):
    # df1 and df2 - tables where search match
    # sheet1 and sheet2 - sheets of tables where search match
    # column1 and column2 - columns between which search match

    # dict with connection indexes from both tables
    width1 = len(df1.columns)
    width2 = len(df2.columns)

    connection_dict = find_connections(df1[column1], df2[column2], final_score)

    # connection dicts with total scores for both tables
    total_connection_1 = {}
    total_connection_2 = {}

    # open both table and create columns
    wb1 = openpyxl.load_workbook(filename=name_table_1, read_only=False)
    ws1 = wb1[sheet1]
    ws1.cell(row=1, column=width1 + 1).value = 'connection_n'
    ws1.cell(row=1, column=width1 + 2).value = 'score_n'

    wb2 = openpyxl.load_workbook(filename=name_table_2, read_only=False)
    ws2 = wb2[sheet2]
    ws2.cell(row=1, column=width2 + 1).value = 'connection_n'
    ws2.cell(row=1, column=width2 + 2).value = 'score_n'

    # calculate and write score and match
    for i, j in connection_dict.items():
        # total_connection_1[i] = []
        # total_connection_2[i] = []
        # if row i from 1 table has match with row j form 2 table:
        if len(j) > 0:
            if len(j) == 1:
                # if one row has 2 mathces with another table rows, it writes as table_name_row1_row2
                if ws1.cell(row=i + 2, column=width1 + 1).value is None:
                    ws1.cell(row=i + 2, column=width1 + 1).value = f'{name_table_2}_{str(j[0][0] + 2)}'
                    total_connection_1[i] = [[j[0][0], j[0][1] * columntype_score_1 / 100]]

                else:
                    ws1.cell(row=i + 2, column=width1 + 1).value = ws1.cell(row=i + 2, column=width1 + 1).value + '_' + str(j[0][0] + 2)
                    total_connection_1[i].append([j[0][0], j[0][1] * columntype_score_1 / 100])

                if ws2.cell(row=j[0][0] + 2, column=width2 + 1).value is None:
                    ws2.cell(row=j[0][0] + 2, column=width2 + 1).value = f'{name_table_1}_{str(i + 2)}'
                    total_connection_2[j[0][0]] = [[i, j[0][1] * columntype_score_2 / 100]]

                else:
                    ws2.cell(row=j[0][0] + 2, column=width2 + 1).value = ws2.cell(row=j[0][0] + 2, column=width2 + 1).value + '_' + str(i + 2)
                    total_connection_2[j[0][0]].append([i, j[0][1] * columntype_score_2 / 100])

                # write score
                ws1.cell(row=i + 2, column=width1 + 2).value = str(j[0][1]*columntype_score_1/100) + ' %'        # name temporary connection * score of column type
                ws2.cell(row=j[0][0] + 2, column=width2 + 2).value = str(j[0][1]*columntype_score_2/100) + ' %'       # name temporary connection * score of column type

            # if for i row there are multiple rows form 2 table
            else:
                for pair in j:
                    if ws1.cell(row=i + 2, column=width1 + 1).value is None:
                        ws1.cell(row=i + 2, column=width1 + 1).value = f'{name_table_2}_{str(pair[0] + 2)}'
                        total_connection_1[i] = [[pair[0], pair[1] * columntype_score_1 / 100]]

                    else:
                        ws1.cell(row=i + 2, column=width1 + 1).value = ws1.cell(row=i + 2, column=width1 + 1).value + '_' + str(pair[0] + 2)
                        total_connection_1[i].append([pair[0], pair[1] * columntype_score_1 / 100])

                    if ws2.cell(row=pair[0] + 2, column=width2 + 1).value is None:
                        ws2.cell(row=pair[0] + 2, column=width2 + 1).value = f'{name_table_1}_{str(i + 2)}'
                        total_connection_2[pair[0]] = [[i, pair[1] * columntype_score_2 / 100]]

                    else:
                        ws2.cell(row=pair[0] + 2, column=width2 + 1).value = ws2.cell(row=pair[0] + 2, column=width2 + 1).value + '_' + str(i + 2)
                        total_connection_2[pair[0]].append([i, pair[1] * columntype_score_2 / 100])

                    # write score
                    ws1.cell(row=i + 2, column=width1 + 2).value = str(pair[1]*columntype_score_1/100) + ' %'         # name temporary connection * score of column type
                    ws2.cell(row=pair[0] + 2, column=width2 + 2).value = str(pair[1]*columntype_score_2/100) + ' %'    # name temporary connection * score of column type

    wb1.save(name_table_1)
    wb2.save(name_table_2)
    return total_connection_1, total_connection_2

# write_connections(df1, sheet_name_1, df1_column, df2, sheet_name_2, df2_column)