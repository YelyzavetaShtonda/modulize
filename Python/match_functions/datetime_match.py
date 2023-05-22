import pandas as pd
import re
import openpyxl
import pymysql
import sys
sys.path.insert(0, 'D:\Робота\Python\Time_parsing')
import Time_parsing.time_and_date_recognition as time_and_date_recognition

pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)

# connect to sql cloud
conn = pymysql.connect(
    host="34.72.200.35",
    database="categorization_db",
    password='R6gK%L!J6034',
    user="root",
    port=3306,
    ssl_disabled=True
)

with conn.cursor() as cursor:
    # Read a single record
    sql = "SELECT * FROM Scoring"
    cursor.execute(sql, ())
    result = cursor.fetchone()
    scoring_df = pd.read_sql_query(sql, conn)

#read excel file
# name_table_1 = 'PSP_to_Airtable_1.xlsx'
# name_table_2 = 'PSP_to_Airtable_1_copy.xlsx'
# sheet_name_1 = 'Volt'
# sheet_name_2 = 'Volt'
# df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
# df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
# df1_column = 'created_at'
# df2_column = 'created_at'
# width1 = len(df1.columns)
# width2 = len(df2.columns)
# length1 = len(df1.index)
# length2 = len(df2.index)

timestamp_scoring = scoring_df.query('`Function name` == "datetimeconnection"').reset_index(drop=True)

# function which calculates the score
def calc_score(data1, data2):
    # takes to lists of parsed datetime and calculate score

    #define score for every type of connection
    entire_match = timestamp_scoring.query('`scoring API name` == "datetimematch 1"')['score'][0]
    year_month_day_minute = timestamp_scoring.query('`scoring API name` == "datetimematch 2"')['score'][1]
    year_month_day_hour_up_to_5_mins = timestamp_scoring.query('`scoring API name` == "datetimematch 3"')['score'][2]
    year_month_day_up_to_5_mins = timestamp_scoring.query('`scoring API name` == "datetimematch 4"')['score'][3]
    no_match = timestamp_scoring.query('`scoring API name` == "datetimematch 5"')['score'][4]

    score = no_match   # 0

    mins_and_secs1 = int(data1[4]) * 60 + int(data1[5])
    mins_and_secs2 = int(data2[4]) * 60 + int(data2[5])

    # If there is a match by year, date, month, day, hours, minute, or second - score 100%
    for i in range(7):
        if data1[i] == data2[i]:
            score = entire_match     # 100
            continue
        else:
            score = no_match         # 0
            break
    if score == entire_match:        # if score == 100:
        return score

    # If there is a match by year, date, month, day, minute - score 95%
    elif data1[0] == data2[0] and data1[1] == data2[1] and data1[2] == data2[2] and data1[4] == data2[4]:
        score = year_month_day_minute      # 95

    # if there is a match by year, date, month, day, hours, with a different of up to 5 minutes - 90%
    elif data1[0] == data2[0] and data1[1] == data2[1] and data1[2] == data2[2] and data1[3] == data2[3] \
                                                     and abs(mins_and_secs1-mins_and_secs2) <= 300:
        score = year_month_day_hour_up_to_5_mins     # 90

    # if there is a match by year, date, month, day, with a difference of up to 5 minutes - 70%
    elif data1[0] == data2[0] and data1[1] == data2[1] and data1[2] == data2[2] \
                                                     and abs(mins_and_secs1-mins_and_secs2) <= 300:
        score = year_month_day_up_to_5_mins     # 70

    return score



# function which look for connection between two columns (take 2 columns and return dictionary with indexes of rows which have connection)
def find_connections(column1, column2, final_score):
    connection_indexes = {}
    for i in range(max(len(column1), len(column2))):
        connection_indexes[i] = []   # for each key we create list of values
        for j in range(len(column2)):
            date1 = time_and_date_recognition.parse_date(column1[i])   # parsed cell from 1 table
            date2 = time_and_date_recognition.parse_date(column2[j])   # parsed cell from 2 table
            score = calc_score(date1, date2)
            # if score > 0 add indexes of rows form tables as key and value to dict
            if score == 0:
                continue
            else:
                connection_indexes[i].append([j, score*final_score/100])  # score of date time function * a score 50%

    # if one key has multiple values (to 1 row in 1 table we have 2+ connections from 2 table) it choose connection with highest score
    for key in connection_indexes.keys():  # loop through all the keys
        value = connection_indexes[key]

        #цей код для того, щоб якщо є метч до 1 рядка з 1 таблиці з багатьма рядками з 2 таблиці, аоно лишало метчі лише з найбільшими скорами (напр якщо є два метчі зі скорами 90б і один з 70, то метч зі скором 70 воно викине)
        # if len(value) > 1:
        #     scores = []
        #     for k in range(len(value)):
        #         scores.append(value[k][1])
        #     connection_indexes[key] = []
        #     for k in range(len(value)):
        #         if value[k][1] == max(scores):
        #             connection_indexes[key].append([value[k][0], value[k][1]])

    return connection_indexes


# function which write connections and scores to new columns to both tables
def write_connections(name_table_1, df1, sheet1, column1, name_table_2, df2, sheet2, column2, final_score, columntype_score_1, columntype_score_2):
    """
    :param df1: tables where search match
    :param sheet1: sheets of tables where search match
    :param column1: columns between which search match
    :param df2: tables where search match
    :param sheet2: sheets of tables where search match
    :param column2: columns between which search match
    :return: None
    """

    width1 = len(df1.columns)
    width2 = len(df2.columns)

    # dict with connection indexes from both tables
    connection_dict = find_connections(df1[column1], df2[column2], final_score)

    # connection dicts with total scores for both tables
    total_connection_1 = {}
    total_connection_2 = {}

    # open both table and create columns
    wb1 = openpyxl.load_workbook(filename=name_table_1, read_only=False)
    ws1 = wb1[sheet1]
    ws1.cell(row=1, column=width1 + 1).value = 'connection_d'
    ws1.cell(row=1, column=width1 + 2).value = 'score_d'

    wb2 = openpyxl.load_workbook(filename=name_table_2, read_only=False)
    ws2 = wb2[sheet2]
    ws2.cell(row=1, column=width2 + 1).value = 'connection_d'
    ws2.cell(row=1, column=width2 + 2).value = 'score_d'


    # calculate and write score and match
    # for i, j in connection_dict.items():
    #     # if row i from 1 table has match with row j form 2 table:
    #     if len(j) > 0:
    #         ws1.cell(row=i + 2, column=width1 + 1).value = f'{name_table_2}_{str(j[0][0] + 2)}'
    #         ws2.cell(row=j[0][0] + 2, column=width2 + 1).value = f'{name_table_1}_{str(i + 2)}'
    #         # write score
    #         ws1.cell(row=i + 2, column=width1 + 2).value = str(j[0][1]) + ' %'
    #         ws2.cell(row=j[0][0] + 2, column=width2 + 2).value = str(j[0][1]) + ' %'

    for i, j in connection_dict.items():
        # total_connection_1[i] = []
        # total_connection_2[i] = []
        # if row i from 1 table has match with row j form 2 table:
        if len(j) > 0:
            if len(j) == 1:
                # if one row has 2 mathces with another table rows, it writes as table_name_row1_row2
                if ws1.cell(row=i + 2, column=width1 + 1).value is None:
                    ws1.cell(row=i + 2, column=width1 + 1).value = f'{name_table_2}_{str(j[0][0] + 2)}'
                    total_connection_1[i] = [[j[0][0], j[0][1] * columntype_score_1 / 100]]
                else:
                    ws1.cell(row=i + 2, column=width1 + 1).value = ws1.cell(row=i + 2, column=width1 + 1).value + '_' + str(j[0][0] + 2)
                    total_connection_1[i].append([j[0][0], j[0][1] * columntype_score_1 / 100])

                if ws2.cell(row=j[0][0] + 2, column=width2 + 1).value is None:
                    ws2.cell(row=j[0][0] + 2, column=width2 + 1).value = f'{name_table_1}_{str(i + 2)}'
                    total_connection_2[j[0][0]] = [[i, j[0][1] * columntype_score_2 / 100]]

                else:
                    ws2.cell(row=j[0][0] + 2, column=width2 + 1).value = ws2.cell(row=j[0][0] + 2, column=width2 + 1).value + '_' + str(i + 2)
                    total_connection_2[j[0][0]].append([i, j[0][1] * columntype_score_2 / 100])

                # write score
                ws1.cell(row=i + 2, column=width1 + 2).value = str(j[0][1]*columntype_score_1/100) + ' %'    # score of date time function * a score 50% * score of column type
                ws2.cell(row=j[0][0] + 2, column=width2 + 2).value = str(j[0][1]*columntype_score_2/100) + ' %'    # score of date time function * a score 50% * score of column type

            # if for i row there are multiple rows form 2 table
            else:
                for pair in j:
                    if ws1.cell(row=i + 2, column=width1 + 1).value is None:
                        ws1.cell(row=i + 2, column=width1 + 1).value = f'{name_table_2}_{str(pair[0] + 2)}'
                        total_connection_1[i] = [[pair[0], pair[1] * columntype_score_1 / 100]]

                    else:
                        ws1.cell(row=i + 2, column=width1 + 1).value = ws1.cell(row=i + 2, column=width1 + 1).value + '_' + str(pair[0] + 2)
                        total_connection_1[i].append([pair[0], pair[1] * columntype_score_1 / 100])


                    if ws2.cell(row=pair[0] + 2, column=width2 + 1).value is None:
                        ws2.cell(row=pair[0] + 2, column=width2 + 1).value = f'{name_table_1}_{str(i + 2)}'
                        total_connection_2[pair[0]] = [[i, pair[1] * columntype_score_2 / 100]]

                    else:
                        ws2.cell(row=pair[0] + 2, column=width2 + 1).value = ws2.cell(row=pair[0] + 2, column=width2 + 1).value + '_' + str(i + 2)
                        total_connection_2[pair[0]].append([i, pair[1] * columntype_score_2 / 100])

                    # write score
                    ws1.cell(row=i + 2, column=width1 + 2).value = str(pair[1]*columntype_score_1/100) + ' %'   # score of date time function * a score 50% * score of column type
                    ws2.cell(row=pair[0] + 2, column=width2 + 2).value = str(pair[1]*columntype_score_2/100) + ' %'   # score of date time function * a score 50% * score of column type

    wb1.save(name_table_1)
    wb2.save(name_table_2)
    return total_connection_1, total_connection_2


# write score and record id to both tables
# write_connections(df1, sheet_name_1, df1_column, df2, sheet_name_2, df1_column)


