from datetime import datetime



import pandas as pd
import re
import openpyxl
import pymysql
import sys
sys.path.insert(0, 'D:\Робота\Python\data_recognition')
import data_recognition.func_to_identify_column_titles as func_to_identify_column_titles
import datetime_match
import transaction_id_match
import email_match
import name_match
import other_type_match
import other_type_contain_match



def find_trx_id(column1_type, column2_type):
    """
    choose what columns are trx_id based on categorize_columns() final_output func for each table
    :return:
    """
    # find trx ids in first table
    list_with_ids_1 = []
    for i in range(len(column1_type.index)):

        if re.search('id', column1_type.loc[i, 'Final output']) or \
           re.search('ID', column1_type.loc[i, 'Final output']) or \
           re.search('Id', column1_type.loc[i, 'Final output']):
                list_with_ids_1.append([column1_type.loc[i, 'Column Name'], column1_type.loc[i, 'Score']])

    # find trx ids in second table
    list_with_ids_2 = []
    for i in range(len(column2_type.index)):

        if re.search('id', column2_type.loc[i, 'Final output']) or \
           re.search('ID', column2_type.loc[i, 'Final output']) or \
           re.search('Id', column2_type.loc[i, 'Final output']):
                list_with_ids_2.append([column2_type.loc[i, 'Column Name'], column2_type.loc[i, 'Score']])
    # print('here', list_with_ids_1, list_with_ids_2)
    return list_with_ids_1, list_with_ids_2

def find_email(column1_type, column2_type):
    # find emails in first table
    list_with_emails_1 = []
    for i in range(len(column1_type.index)):

        if re.search('email', column1_type.loc[i, 'Final output']) or re.search('Email', column1_type.loc[i, 'Final output']):
            list_with_emails_1.append([column1_type.loc[i, 'Column Name'], column1_type.loc[i, 'Score']])
    # print(list_with_emails_1)

    # find emails in second table
    list_with_emails_2 = []
    for i in range(len(column2_type.index)):

        if re.search('email', column2_type.loc[i, 'Final output']) or re.search('Email', column2_type.loc[i, 'Final output']) or re.search('Id', column2_type.loc[i, 'Final output']):
            list_with_emails_2.append([column2_type.loc[i, 'Column Name'], column2_type.loc[i, 'Score']])
    # print('her', list_with_emails_1 , list_with_emails_2)
    return list_with_emails_1, list_with_emails_2

def find_names(column1_type, column2_type):
    # find names in first table
    list_with_names_1 = []
    # print(column1_type)
    for i in range(len(column1_type.index)):

        if re.search('Cardholder full name', column1_type.loc[i, 'Final output']) \
            or re.search('FIRST NAME', column1_type.loc[i, 'Final output']) \
            or re.search('LAST NAME', column1_type.loc[i, 'Final output']) \
            or re.search('Full name', column1_type.loc[i, 'Final output']):

            list_with_names_1.append([column1_type.loc[i, 'Column Name'], column1_type.loc[i, 'Score']])
    # print(list_with_names_1)

    # find names in second table
    list_with_names_2 = []
    for i in range(len(column2_type.index)):

        if  re.search('Cardholder full name', column2_type.loc[i, 'Final output']) \
            or re.search('FIRST NAME', column2_type.loc[i, 'Final output']) \
            or re.search('LAST NAME', column2_type.loc[i, 'Final output']) \
            or re.search('Full name', column2_type.loc[i, 'Final output']):

            list_with_names_2.append([column2_type.loc[i, 'Column Name'], column2_type.loc[i, 'Score']])
    # print(list_with_names_2)
    return list_with_names_1, list_with_names_2

def find_datetime(column1_type, column2_type):
    # find datetime in first table
    list_with_datetime_1 = []
    for i in range(len(column1_type.index)):

        if re.search('date', column1_type.loc[i, 'Final output']) \
                or re.search('Date', column1_type.loc[i, 'Final output']) \
                or re.search('DATE', column1_type.loc[i, 'Final output']) \
                or re.search('time', column1_type.loc[i, 'Final output'])\
                or re.search('Time', column1_type.loc[i, 'Final output'])\
                or re.search('TIME', column1_type.loc[i, 'Final output']):

            list_with_datetime_1.append([column1_type.loc[i, 'Column Name'], column1_type.loc[i, 'Score']])
    # print(list_with_datetime_1)

    # find datetime in second table
    list_with_datetime_2 = []
    for i in range(len(column2_type.index)):

        if re.search('date', column2_type.loc[i, 'Final output']) \
                or re.search('Date', column2_type.loc[i, 'Final output']) \
                or re.search('DATE', column2_type.loc[i, 'Final output']) \
                or re.search('time', column2_type.loc[i, 'Final output'])\
                or re.search('Time', column2_type.loc[i, 'Final output'])\
                or re.search('TIME', column2_type.loc[i, 'Final output']):

            list_with_datetime_2.append([column2_type.loc[i, 'Column Name'], column2_type.loc[i, 'Score']])
    # print('here', list_with_datetime_1, list_with_datetime_2)
    return list_with_datetime_1, list_with_datetime_2

def find_other_types(lst1, lst2, column1_type, column2_type):
    """
    :param lst1 and lst2: lists with columns names and scores that are recognized as id, email, name, or datetime for each table
    :return: array of lists with columns names and scores that are NOT recognized as id, email, name, or datetime for each table
    """
    list_1 = []
    list_2 = []

    dataframe_1 = column1_type[['Column Name', 'Score']]
    dataframe_2 = column2_type[['Column Name', 'Score']]

    def parse_list(lst):
        # takes list of lists of name and score and return list with names
        names = []
        for name, score in lst:
            names.append(name)
        return names

    lst_1 = parse_list(lst1)
    lst_2 = parse_list(lst2)

    for i in range(len(dataframe_1.index)):
        if not dataframe_1.loc[i, 'Column Name'] in lst_1:
            list_1.append([dataframe_1.loc[i, 'Column Name'], dataframe_1.loc[i, 'Score']])

    for i in range(len(dataframe_2.index)):
        if not dataframe_2.loc[i, 'Column Name'] in lst_2:
            list_2.append([dataframe_2.loc[i, 'Column Name'], dataframe_2.loc[i, 'Score']])

    return list_1, list_2

def find_other_contain_types(lst1, lst2, column1_type, column2_type):
    """
        :param lst1 and lst2: lists with columns names and scores that are recognized as id, email, name, or datetime for each table
        :return: array of lists with columns names and scores that are NOT recognized as id, email, name, or datetime for each table
        """
    list_1 = []
    list_2 = []

    dataframe_1 = column1_type[['Column Name', 'Score']]
    dataframe_2 = column2_type[['Column Name', 'Score']]

    def parse_list(lst):
        # takes list of lists of name and score and return list with names
        names = []
        for name, score in lst:
            names.append(name)
        return names

    lst_1 = parse_list(lst1)
    lst_2 = parse_list(lst2)

    for i in range(len(dataframe_1.index)):
        if not dataframe_1.loc[i, 'Column Name'] in lst_1:
            list_1.append([dataframe_1.loc[i, 'Column Name'], dataframe_1.loc[i, 'Score']])

    for i in range(len(dataframe_2.index)):
        if not dataframe_2.loc[i, 'Column Name'] in lst_2:
            list_2.append([dataframe_2.loc[i, 'Column Name'], dataframe_2.loc[i, 'Score']])

    return list_1, list_2

def write_add_trx_id_scores(i, matches_number_i, j, matches_number_j, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2):
    # receives index of row and number of matches with this row from 1st table. receives the same from another table
    # print(f'INCREASE SCORE BY {5*(matches_number_i-1)} and {5*(matches_number_j-1)} IN {i+2} ~ {j+2} rows')
    print('i`m here')
    wb1 = openpyxl.load_workbook(filename=name_table_1, read_only=False)
    ws1 = wb1[sheet_name_1]
    ws1.cell(row=i + 2, column=width1 + 2).value = str(float(ws1.cell(row=i+2, column=width1 + 2).value[:-2]) + 5*(matches_number_i-1)) + ' %'


    wb2 = openpyxl.load_workbook(filename=name_table_2, read_only=False)
    ws2 = wb2[sheet_name_2]
    ws2.cell(row=j+2, column=width2 + 2).value = str(float(ws2.cell(row=j+2, column=width2 + 2).value[:-2]) + 5*(matches_number_j-1)) + ' %'

    wb1.save(name_table_1)
    wb2.save(name_table_2)

def additional_trx_id_scores(lists, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2, length1, length2, df1, df2):
    # receives lists of columns and scores, which needed to be searched for additional scores
    print('I`m here')
    list_1, list_2 = lists

    # number of matches in every i row form 1 table
    matches_in_every_row = {}
    # по кожній колонці 1 таблиці
    for col1, score1 in list_1:
        # по кожному рядку колонки з 1 таблиці
        for i in range(length1):
            # по кожному рядку 2 таблиці (через всі колонки)
            for j in range(length2):
                matches_number = 0
                # по кожній колонці з рядка в 2 таблиці
                for col2, score2 in list_2:
                    # have 2 columns from two tables. for every record from 1st column search matches with every row in 2nd column
                    # print(col1, '***', col2)
                    if df1[col1][i] == df2[col2][j]:
                        matches_number += 1
                # print(f'change_scores (found {matches_number} matches {i+2} row with {j+2} row)')
                # add match number per row in dict
                if i in matches_in_every_row:
                    matches_in_every_row[i] += matches_number
                else:
                    matches_in_every_row[i] = matches_number
                if matches_number > 1:
                    write_add_trx_id_scores(i, matches_in_every_row[i], j, matches_number, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2)


# function tries match by trx id, if no then by email, if no then by names, if no then by datetime
def final_scores(priority, final_connection_temp, column1_type, column2_type, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2, length1, length2):

    # global df1, df2

    return_list = []

    other_column_types_list_1 = []   # list with names of columns that are not ids, emails, names or datetime in 1 table
    other_column_types_list_2 = []   # list with names of columns that are not ids, emails, names or datetime in 1 table

    def write_columns_in_list(array_):
        for i in array_[0]:
            other_column_types_list_1.append(i)
        for i in array_[1]:
            other_column_types_list_2.append(i)

    def trx_match_func():
        # lists with column names which are recognized as trx_id type
        ids = datetime.now()
        trx_id_lists = find_trx_id(column1_type, column2_type)
        # write_columns_in_list(trx_id_lists)  # write column names od id columns to lists
        if len(trx_id_lists[0]) and len(trx_id_lists[1]):
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = trx_id_lists[0][0][0]
            column_name_2 = trx_id_lists[1][0][0]
            columntype_score_1 = trx_id_lists[0][0][1]
            columntype_score_2 = trx_id_lists[1][0][1]
            # print('by trx')
            final_trx_score = final_connection_temp.query('`scoring API name` == "TRX ID temporary connection"')['score'][0]
            # list of dicts with matches and final scores per every match type (by trx_id, email, name, and datetime)
            final_trx_id_matches = transaction_id_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_trx_score, columntype_score_1, columntype_score_2)
            find_id_in_dfs = datetime.now()
            print('find_id_in_dfs', find_id_in_dfs - ids)
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            additional_trx_id_scores(trx_id_lists, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2, length1, length2, df1, df2)
            count_add_temp_score_id = datetime.now()
            print('count_add_temp_score_id', count_add_temp_score_id-find_id_in_dfs)
            return_list.append(['TRX_ID', final_trx_id_matches])

    def email_match_func():
        # lists with column names which are recognized as email type
        emails = datetime.now()
        email_lists = find_email(column1_type, column2_type)
        # write_columns_in_list(email_lists)  # write column names od id columns to lists
        if len(email_lists[0]) > 0 and len(email_lists[1]) > 0:
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = email_lists[0][0][0]
            column_name_2 = email_lists[1][0][0]
            columntype_score_1 = email_lists[0][0][1]
            columntype_score_2 = email_lists[1][0][1]
            # print('by email')
            # print(email_lists)
            final_email_score = final_connection_temp.query('`scoring API name` == "email temporary connection"')['score'][1]
            # list of dicts with matches and final scores per every match type (by trx_id, email, name, and datetime)
            final_email_matches = email_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_email_score, columntype_score_1, columntype_score_2)
            # find_and_write_match_email = datetime.now()
            # print('find_and_write_match_email',find_and_write_match_email-count_add_temp_score_id)
            # additional_trx_id_scores(email_lists)
            # count_add_temp_score_email = datetime.now()
            # print('count_add_temp_score_email', count_add_temp_score_email-find_and_write_match_email)
            # print(final_email_matches[0])
            # print(final_email_matches[1])
            return_list.append(['email', final_email_matches])
            find_email_in_dfs = datetime.now()
            print('find_email_in_dfs', find_email_in_dfs-emails)

    def name_match_func():
        # lists with column names which are recognized as name type
        names=datetime.now()
        name_lists = find_names(column1_type, column2_type)
        # write_columns_in_list(name_lists)  # write column names od id columns to lists
        if len(name_lists[0]) > 0 and len(name_lists[1]) > 0:
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = name_lists[0][0][0]
            column_name_2 = name_lists[1][0][0]
            columntype_score_1 = name_lists[0][0][1]
            columntype_score_2 = name_lists[1][0][1]
            # print('by name')
            # print(name_lists)
            final_name_score = final_connection_temp.query('`scoring API name` == "name temporary connection"')['score'][2]
            # list of dicts with matches and final scores per every match type (by trx_id, email, name, and datetime)
            final_name_matches = name_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_name_score, columntype_score_1, columntype_score_2)
            # print(final_name_matches[0])
            # print(final_name_matches[1])
            return_list.append(['name', final_name_matches])
            find_names_in_dfs = datetime.now()
            print('find_names_in_dfs', find_names_in_dfs-names)

    def datetime_match_func():
        # lists with column names which are recognized as datetime type
        datetimes = datetime.now()
        datetime_lists = find_datetime(column1_type, column2_type)
        # write_columns_in_list(datetime_lists)  # write column names od id columns to lists
        if len(datetime_lists[0]) > 0 and len(datetime_lists[1]) > 0:
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = datetime_lists[0][0][0]
            column_name_2 = datetime_lists[1][0][0]
            columntype_score_1 = datetime_lists[0][0][1]
            columntype_score_2 = datetime_lists[1][0][1]
            # print('by datetime')
            # print(datetime_lists)
            final_datetime_score = final_connection_temp.query('`scoring API name` == "date temporary connection"')['score'][3]
            # list of dicts with matches and final scores per every match type (by trx_id, email, name, and datetime)
            final_datetime_matches = datetime_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_datetime_score, columntype_score_1, columntype_score_2)
            # print(final_datetime_matches[0])
            # print(final_datetime_matches[1])
            return_list.append(['datetime', final_datetime_matches])
            find_datetime_in_dfs = datetime.now()
            print('find_datetime_in_dfs', find_datetime_in_dfs-datetimes)

    def other_match_func():
        # defines column that are id, emails, names or datetimes
        trx_id_lists = find_trx_id(column1_type, column2_type)
        write_columns_in_list(trx_id_lists)
        email_lists = find_email(column1_type, column2_type)
        write_columns_in_list(email_lists)
        name_lists = find_names(column1_type, column2_type)
        write_columns_in_list(name_lists)
        datetime_lists = find_datetime(column1_type, column2_type)
        write_columns_in_list(datetime_lists)

        # lists with column names which are recognized as other type
        others = datetime.now()
        other_type_lists = find_other_types(other_column_types_list_1, other_column_types_list_2, column1_type, column2_type)
        # print('other types ', other_type_lists[0])
        # print(other_type_lists[1])
        if len(other_type_lists[0]) and len(other_type_lists[1]):
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = other_type_lists[0][0][0]
            column_name_2 = other_type_lists[1][0][0]
            columntype_score_1 = other_type_lists[0][0][1]
            columntype_score_2 = other_type_lists[1][0][1]
            final_other_score = final_connection_temp.query('`scoring API name` == "noname temporary connection"')['score'][4]
            final_other_matches = other_type_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_other_score, columntype_score_1, columntype_score_2)
            return_list.append(['other', final_other_matches])
            find_other_in_dfs = datetime.now()
            print('find_other_matches_in_dfs', find_other_in_dfs - others)

    def other_contain_match_func():
        # # defines column that are id, emails, names or datetimes
        # trx_id_lists = find_trx_id()
        # write_columns_in_list(trx_id_lists)
        # email_lists = find_email()
        # write_columns_in_list(email_lists)
        # name_lists = find_names()
        # write_columns_in_list(name_lists)
        # datetime_lists = find_datetime()
        # write_columns_in_list(datetime_lists)

        # lists with column names which are recognized as other type

        others = datetime.now()
        other_type_lists = find_other_contain_types(other_column_types_list_1, other_column_types_list_2, column1_type, column2_type)
        if len(other_type_lists[0]) and len(other_type_lists[1]):
            df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
            df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
            column_name_1 = other_type_lists[0][0][0]
            column_name_2 = other_type_lists[1][0][0]
            columntype_score_1 = other_type_lists[0][0][1]
            columntype_score_2 = other_type_lists[1][0][1]
            print(column_name_1, column_name_2)
            final_other_score = final_connection_temp.query('`scoring API name` == "noname temporary connection"')['score'][4]
            final_other_matches = other_type_contain_match.write_connections(name_table_1, df1, sheet_name_1, column_name_1, name_table_2, df2, sheet_name_2, column_name_2, final_other_score, columntype_score_1, columntype_score_2)
            return_list.append(['other_contain', final_other_matches])
            find_other_in_dfs = datetime.now()
            print('find_other_contain_matches_in_dfs', find_other_in_dfs - others)

    # call functions to search for matches by priority
    for i in range(len(priority)):

        if priority[i] == 'TRX_ID':
            trx_match_func()
        elif priority[i] == 'email':
            email_match_func()
        elif priority[i] == 'name':
            name_match_func()
        elif priority[i] == 'datetime':
            datetime_match_func()
        elif priority[i] == 'other':
            other_match_func()
        elif priority[i] == 'other_contain':
            other_contain_match_func()


    return return_list


def write_final_score(dicts, final_coeff, name_table_1, name_table_2, sheet_name_1, sheet_name_2):
    """
    :param dicts: list of dictionaries with matches and final scores for both tables
    :param final_coeff: coefficient to be multiplied by existing score to receive final score
    :return: writes final scores to both tables
    """
    dict_1, dict_2 = dicts

    # import dataframes to find out the width
    df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
    df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
    width1 = len(df1.columns)
    width2 = len(df2.columns)

    # open 1 table
    wb1 = openpyxl.load_workbook(filename=name_table_1, read_only=False)
    ws1 = wb1[sheet_name_1]
    ws1.cell(row=1, column=width1 + 1).value = 'final_connection'
    ws1.cell(row=1, column=width1 + 2).value = 'final_score'
    # open 2 table
    wb2 = openpyxl.load_workbook(filename=name_table_2, read_only=False)
    ws2 = wb2[sheet_name_2]
    ws2.cell(row=1, column=width2 + 1).value = 'final_connection'
    ws2.cell(row=1, column=width2 + 2).value = 'final_score'


    # print('dict1', dict_1)
    # print('dict2', dict_2)

    #write matches ans final scores to 1st table
    for row, score in dict_1.items():
        if len(score) > 0:
            if len(score) == 1:
                if ws1.cell(row=row + 2, column=width1 + 1).value is None:
                    ws1.cell(row=row + 2, column=width1 + 1).value = f'{name_table_2}_{str(score[0][0] + 2)}'
                else:
                    ws1.cell(row=row + 2, column=width1 + 1).value = ws1.cell(row=row + 2, column=width1 + 1).value + '_' + str(score[0][0] + 2)
               # write score
                ws1.cell(row=row + 2, column=width1 + 2).value = str(score[0][1]*final_coeff/100) + ' %'

            # if for i row there are multiple matches with rows from 2 table
            else:
                for pair in score:
                    if ws1.cell(row=row + 2, column=width1 + 1).value is None:
                        ws1.cell(row=row + 2, column=width1 + 1).value = f'{name_table_2}_{str(pair[0] + 2)}'
                    else:
                        ws1.cell(row=row + 2, column=width1 + 1).value = ws1.cell(row=row + 2, column=width1 + 1).value + '_' + str(pair[0] + 2)

                    # write score
                    ws1.cell(row=row + 2, column=width1 + 2).value = str(pair[1] * final_coeff / 100) + ' %'

    # write matches ans final scores to 1st table
    for row, score in dict_2.items():
        if len(score) > 0:
            if len(score) == 1:
                if ws2.cell(row=row + 2, column=width2 + 1).value is None:
                    ws2.cell(row=row + 2, column=width2 + 1).value = f'{name_table_2}_{str(score[0][0] + 2)}'
                else:
                    ws2.cell(row=row + 2, column=width2 + 1).value = ws2.cell(row=row + 2, column=width2 + 1).value + '_' + str(score[0][0] + 2)
               # write score
                ws2.cell(row=row + 2, column=width2 + 2).value = str(score[0][1]*final_coeff/100) + ' %'

            # if for i row there are multiple matches with rows from 2 table
            else:
                for pair in score:
                    if ws2.cell(row=row + 2, column=width2 + 1).value is None:
                        ws2.cell(row=row + 2, column=width2 + 1).value = f'{name_table_2}_{str(pair[0] + 2)}'
                    else:
                        ws2.cell(row=row + 2, column=width2 + 1).value = ws2.cell(row=row + 2, column=width2 + 1).value + '_' + str(pair[0] + 2)

                    # write score
                    ws2.cell(row=row + 2, column=width2 + 2).value = str(pair[1] * final_coeff / 100) + ' %'



    wb1.save(name_table_1)
    wb2.save(name_table_2)


def final_scores_2(list_of_dicts, priority, name_table_1, name_table_2, sheet_name_1, sheet_name_2):
    """
    takes list of dicts with scores for each type and calculate final scores to pass it to write_final_score() func
    """
    time1 = datetime.now()
    # list of dicts with matches and scores for each type
    trx_id_dict = {}
    email_dict = {}
    name_dict = {}
    datetime_dict = {}

    # list of 4 dicts above in prioritized order
    list_of_prioritized_dicts = []

    # for i in range(len(priority)):
    #
    #     if priority[i] == 'TRX_ID':
    #         trx_match_func()
    #     elif priority[i] == 'email':
    #         email_match_func()
    #     elif priority[i] == 'name':
    #         name_match_func()
    #     elif priority[i] == 'datetime':
    #         datetime_match_func()
    #     else:
    #         other_match_func()

    # parse list with matches&scores of each type
    if len(list_of_dicts) > 0:

        for type, dict in list_of_dicts:
            if type == 'TRX_ID':
                trx_id_dict = dict
                list_of_prioritized_dicts.append([type, dict])
            elif type == 'email':
                email_dict = dict
                list_of_prioritized_dicts.append([type, dict])
            elif type == 'name':
                name_dict = dict
                list_of_prioritized_dicts.append([type, dict])
            elif type == 'datetime':
                datetime_dict = dict
                list_of_prioritized_dicts.append([type, dict])
            elif type == 'other':
                datetime_dict = dict
                list_of_prioritized_dicts.append([type, dict])

        # print('1 list_of_prioritezed dicts', list_of_prioritized_dicts[0])
        # print('2 list_of_prioritezed dicts', list_of_prioritized_dicts[1])

        if priority[0] == list_of_prioritized_dicts[0][0]:
        # if trx_id_dict:

            # if other dicts don`t exist
            if len(list_of_prioritized_dicts) == 1:
            # if not (email_dict or name_dict or datetime_dict):
                write_final_score(list_of_prioritized_dicts[0][1], 100, name_table_1, name_table_2, sheet_name_1, sheet_name_2)
                # write_final_score(trx_id_dict, 100)
            else:
                write_final_score(list_of_prioritized_dicts[0][1], 75, name_table_1, name_table_2, sheet_name_1, sheet_name_2)
                # write_final_score(trx_id_dict, 75)
    write_final_score_ = datetime.now()
    print('write_final_score_', write_final_score_-time1)


def main():
    start = datetime.now()

    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_rows", None)

    # connect to sql cloud
    conn = pymysql.connect(
        host="34.72.200.35",
        database="categorization_db",
        password='R6gK%L!J6034',
        user="root",
        port=3306,
        ssl_disabled=True
    )

    with conn.cursor() as cursor:
        # Read a single record
        sql = "SELECT * FROM Scoring"
        cursor.execute(sql, ())
        result = cursor.fetchone()
        scoring_df = pd.read_sql_query(sql, conn)

    connect_to_sql = datetime.now()
    print('connect to sql ', connect_to_sql - start)


    final_connection = scoring_df.query('`Function name` == "Connection"').reset_index(drop=True)
    final_connection_temp = final_connection.query('`scoring type` == "temporary score"').reset_index(drop=True)
    # final_connection_final = final_connection.query('`scoring type` == "final score"').reset_index(drop=True)

    take_data_from_sql = datetime.now()
    print('take_data_from_sql', take_data_from_sql - connect_to_sql)


    # read excel file
    name_table_1 = 'PSP_to_Airtable_1.xlsx'
    name_table_2 = 'PSP_to_Airtable_1_copy.xlsx'
    sheet_name_1 = 'Fibonatix'
    sheet_name_2 = 'Fibonatix'
    df1 = pd.read_excel(name_table_1, sheet_name=sheet_name_1)
    df2 = pd.read_excel(name_table_2, sheet_name=sheet_name_2)
    width1 = len(df1.columns)
    width2 = len(df2.columns)
    length1 = len(df1.index)
    length2 = len(df2.index)


    read_excel = datetime.now()
    print('read_excel', read_excel - take_data_from_sql)

    column1_type = pd.DataFrame(func_to_identify_column_titles.categorize_columns(name_table_1, sheet_name_1)[0][
                                    ['Column Name', 'Final output',
                                     'Score']])  # таблиця з назвами колонок і final категоризацією
    column2_type = pd.DataFrame(func_to_identify_column_titles.categorize_columns(name_table_2, sheet_name_2)[0][
                                    ['Column Name', 'Final output',
                                     'Score']])  # таблиця з назвами колонок і final категоризацією
    Olyas_func_for_both_tables = datetime.now()
    print('Olyas_func_for_both_tables', Olyas_func_for_both_tables - read_excel)


    priority = ['email', 'name', 'TRX_ID', 'datetime', 'other', 'other_contain']
    list_of_dicts = final_scores(priority, final_connection_temp, column1_type, column2_type, name_table_1, name_table_2, sheet_name_1, sheet_name_2, width1, width2, length1, length2)
    final_scores_2(list_of_dicts, priority, name_table_1, name_table_2, sheet_name_1, sheet_name_2)
    finish = datetime.now()
    print('all time', finish - start)



main()

